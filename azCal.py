from selenium import webdriver
from selenium.webdriver.support.ui import Select
from pprint import pprint
from openpyxl import Workbook, drawing as draw
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
import requests, time


# WorkBook(파일) 생성
wb = Workbook()
ws = wb.active
ws.title = 'Virtual-Machine'

driver = webdriver.Chrome('C:\python\chromedriver')
driver.get("https://azure.microsoft.com/ko-kr/pricing/calculator")

time.sleep(10)


vm = driver.find_element_by_xpath('//*[@id="products-picker-panel"]/div[2]/div[2]/div[1]/div/div/div[1]/button')
vm.click()

time.sleep(5)


region = driver.find_elements_by_css_selector('select[name=region]>option')
operatingSystem = driver.find_elements_by_css_selector('select[name=operatingSystem]>option')

sel_os = Select(driver.find_element_by_xpath('//*[@name="operatingSystem"]'))

col = ['A', 'C']
index = 0

for osTag in operatingSystem :
    os = osTag.get_attribute('value')
    ws[f'{col[index]}1'] = os
    sel_os.select_by_value(os)
    tp = driver.find_elements_by_css_selector('select[name=type]>option')
    
    i = 3
    for t in tp :
        ws[f'{col[index]}{i}'] = t.text
        i += 1
    index += 1
    
wb.save("Azure_Cal.xlsx")





