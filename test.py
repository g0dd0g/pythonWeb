from selenium import webdriver
from selenium.webdriver.support.ui import Select
from pprint import pprint
from openpyxl import Workbook, drawing as draw
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
import requests, time


# WorkBook(파일) 생성
wb = Workbook()
ws = wb.active
ws.title = 'Virtual-Machine'

# Webdriver 사용하여 접속
driver = webdriver.Chrome('C:\python\chromedriver')
driver.get("https://azure.microsoft.com/ko-kr/pricing/calculator")

# 페이지가 로드되는 시간을 위하여 정지
time.sleep(7)

# Virtual-Machine 버튼 클릭하여 추가
vm = driver.find_element_by_xpath('//*[@id="products-picker-panel"]/div[2]/div[2]/div[1]/div/div/div[1]/button')
vm.click()

# VM 추가되는 동안 정지
time.sleep(5)

# 지역 리스트 갖고오기
#region = driver.find_elements_by_css_selector('select[name=region]>option')

currency = ['KRW', 'USD']
cur_text = ['한국 원(₩)', '미국 달러($)']
cur_size = len(currency) 
region = ['korea-central', 'korea-south']
reg_text = ['한국 중부', '한국 남부']
reg_size = len(region)
operatingSystem = driver.find_elements_by_css_selector('select[name=operatingSystem]>option')


sel_region = Select(driver.find_element_by_xpath('//*[@name="region"]'))
sel_currency = Select(driver.find_element_by_xpath('//*[@class="select currency-dropdown"]'))
sel_os = Select(driver.find_element_by_xpath('//*[@name="operatingSystem"]'))
sel_type = Select(driver.find_element_by_xpath('//*[@name="type"]'))
sel_tier = Select(driver.find_element_by_xpath('//*[@name="tier"]'))


col_text = ["통화", "지역", "운영 체제", "유형", "계층", "인스턴스", "인스턴스2", "인스턴스3"]
index = 1

# 머리글 추가
ws['A1'] = col_text[0]
ws['B1'] = col_text[1]
ws['C1'] = col_text[2]
ws['D1'] = col_text[3]
ws['E1'] = col_text[4]
ws['F1'] = col_text[5]

ws['G1'] = col_text[6]
ws['H1'] = col_text[7]


hours = driver.find_element_by_name("hours")
hours.clear()
hours.send_keys("1")

sel_currency.select_by_value(currency[0])
sel_region.select_by_value(region[0])
sel_os.select_by_value('windows')
sel_type.select_by_value('os-only')
sel_tier.select_by_value('standard')

sel_instance = Select(driver.find_element_by_xpath('//*[@name="size"]'))
sel_instance.select_by_value('d8dv4')


savings = driver.find_elements_by_class_name('savings-option')

print(savings[0].find_element_by_class_name('text-heading5').text)
"""
for s in savings :
    print(s.find_element_by_class_name('text-heading5').text)
"""

index += 1
ws[f'A{index}'] = currency[0]
ws[f'B{index}'] = region[0]
ws[f'C{index}'] = 'windows'
ws[f'D{index}'] = 'os-only'
ws[f'E{index}'] = 'standard'

radio_com = savings[0].find_elements_by_css_selector('input')

for rc in radio_com :
    if rc.is_enabled() :
        rc.click()
        ws[f'F{index}'] = driver.find_element_by_css_selector('select[name=size]>option[value=d8dv4]').text
    else :
        ws[f'F{index}'] = "X"
    index += 1          

wb.save("test.xlsx")





